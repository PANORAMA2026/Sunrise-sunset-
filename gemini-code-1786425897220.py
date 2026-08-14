from datetime import datetime, time, timedelta
import io
import math
import re
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
import pandas as pd
import streamlit as st

st.set_page_config(
    page_title="Naval Ephemeris & Automatic Route Mapper", layout="wide"
)

st.title("⚓ Naval Ephemeris & Automatic Route Mapper")
st.write(
    "Strict Sequence Matching & Smooth Sea Day Spatial Interpolation Across Active Route Days."
)

# -----------------------------------------------------------------------------
# STANDARD PORT DATABASE & LOOKUP
# -----------------------------------------------------------------------------
PORT_DB = {
    "LONG BEACH": {"lat": 33.7542, "lon": -118.1982, "tz": "7 W"},
    "LOS ANGELES": {"lat": 33.7542, "lon": -118.1982, "tz": "7 W"},
    "CABO SAN LUCAS": {"lat": 22.8808, "lon": -109.9125, "tz": "7 W"},
    "MAZATLAN": {"lat": 23.2167, "lon": -106.4167, "tz": "7 W"},
    "PUERTO VALLARTA": {"lat": 20.6236, "lon": -105.2325, "tz": "7 W"},
    "ENSENADA": {"lat": 31.8667, "lon": -116.6333, "tz": "7 W"},
    "SAN DIEGO": {"lat": 32.7157, "lon": -117.1611, "tz": "7 W"},
    "CATALINA": {"lat": 33.3428, "lon": -118.3282, "tz": "7 W"},
}


def lookup_port_coords(port_text, db):
    """Search for matching port in the database."""
    if not port_text:
        return None
    pt_upper = str(port_text).upper().strip()

    for key, data in db.items():
        if key in pt_upper or pt_upper in key:
            return data["lat"], data["lon"], data["tz"]
    return None


# -----------------------------------------------------------------------------
# UTILITY FUNCTIONS
# -----------------------------------------------------------------------------


def parse_coordinate(coord):
    """Converts nautical or decimal coordinates to float."""
    if pd.isna(coord):
        return 0.0
    if isinstance(coord, (int, float)):
        return float(coord)

    s = str(coord).strip()
    try:
        return float(s)
    except ValueError:
        pass

    match = re.search(r"(\d+)°\s*(\d+(?:\.\d+)?)\s*['′]?\s*([NSEWnsew]?)", s)
    if match:
        degrees = float(match.group(1))
        minutes = float(match.group(2))
        direction = match.group(3).upper()

        decimal_degrees = degrees + (minutes / 60.0)
        if direction in ["S", "W"]:
            decimal_degrees = -decimal_degrees
        return decimal_degrees

    match_simple = re.search(r"([+-]?\d+(?:\.\d+)?)\s*([NSEWnsew]?)", s)
    if match_simple:
        val = float(match_simple.group(1))
        direction = match_simple.group(2).upper()
        if direction in ["S", "W"]:
            val = -abs(val)
        return val

    return 0.0


def parse_tz_offset(tz_val):
    """Extracts numeric time offset from CSV."""
    if pd.isna(tz_val):
        return 0.0
    s = str(tz_val).strip()
    if not s or s.lower() in ["nan", "none"]:
        return 0.0

    match_hhmmss = re.search(r"([+-]?)\s*(\d{1,2}):(\d{2})(?::(\d{2}))?", s)
    if match_hhmmss:
        sign = -1.0 if match_hhmmss.group(1) == "-" else 1.0
        hours = float(match_hhmmss.group(2))
        minutes = float(match_hhmmss.group(3))
        return sign * (hours + minutes / 60.0)

    match_we = re.search(r"(\d+(?:\.\d+)?)\s*([WEwe])", s)
    if match_we:
        val = float(match_we.group(1))
        direction = match_we.group(2).upper()
        return -val if direction == "W" else val

    try:
        return float(s)
    except ValueError:
        return 0.0


def format_tz_string(tz_val, offset_hours):
    """Formats Time Zone string (e.g., '7 W' or '2 E')."""
    if tz_val and str(tz_val).strip().lower() not in [
        "nan",
        "none",
        "",
        "0.0",
        "0",
    ]:
        s = str(tz_val).strip()
        match_we = re.search(r"(\d+(?:\.\d+)?)\s*([WEwe])", s)
        if match_we:
            return f"{match_we.group(1)} {match_we.group(2).upper()}"

    if offset_hours == 0 or offset_hours is None:
        return "0"

    abs_h = abs(offset_hours)
    h_str = f"{int(abs_h)}" if abs_h.is_integer() else f"{abs_h}"
    direction = "W" if offset_hours < 0 else "E"
    return f"{h_str} {direction}"


def calculate_sun_events_native(lat, lon, date_obj, tz_offset_hours):
    """Calculates sunrise and sunset in local time."""
    try:
        day_of_year = date_obj.timetuple().tm_yday
        declination = 0.409 * math.sin(
            (2 * math.pi / 365) * (day_of_year - 81)
        )
        lat_rad = math.radians(lat)

        cos_ha = (
            math.sin(math.radians(-0.833))
            - math.sin(lat_rad) * math.sin(declination)
        ) / (math.cos(lat_rad) * math.cos(declination))

        if cos_ha >= 1:
            return "Never rises", "Never rises"
        if cos_ha <= -1:
            return "Never sets", "Never sets"

        ha = math.degrees(math.acos(cos_ha))
        b = (2 * math.pi / 365) * (day_of_year - 81)
        eot = 9.87 * math.sin(2 * b) - 7.53 * math.cos(b) - 1.5 * math.sin(b)

        solar_noon_utc = (720 - (4 * lon) - eot) / 60.0

        sunrise_utc_hours = solar_noon_utc - (ha / 15.0)
        sunset_utc_hours = solar_noon_utc + (ha / 15.0)

        sunrise_local_hours = (sunrise_utc_hours + tz_offset_hours) % 24
        sunset_local_hours = (sunset_utc_hours + tz_offset_hours) % 24

        def hours_to_time_str(h_decimal):
            h = int(h_decimal)
            m = int((h_decimal - h) * 60)
            return f"{h:02d}:{m:02d}"

        return hours_to_time_str(sunrise_local_hours), hours_to_time_str(
            sunset_local_hours
        )
    except Exception:
        return "N/A", "N/A"


def clean_text_val(val):
    if pd.isna(val):
        return ""
    s = str(val).strip()
    if s.lower() in ["nan", "none", "null"]:
        return ""
    return s


def export_styled_excel(df_out: pd.DataFrame) -> bytes:
    """Generates styled Excel file with strict row-by-row mapping."""
    df_clean = df_out.dropna(how="all").reset_index(drop=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Schedule"

    ws.views.sheetView[0].showGridLines = True

    font_header = Font(name="Calibri", size=11, bold=True, color="1F4E79")
    font_date = Font(name="Calibri", size=11, bold=True, color="1F4E79")
    font_body = Font(name="Calibri", size=11, bold=True, color="000000")

    fill_header = PatternFill(
        start_color="D0CECE", end_color="D0CECE", fill_type="solid"
    )
    fill_sea_day = PatternFill(
        start_color="BDD7EE", end_color="BDD7EE", fill_type="solid"
    )
    fill_long_beach = PatternFill(
        start_color="A9D08E", end_color="A9D08E", fill_type="solid"
    )
    fill_white = PatternFill(
        start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"
    )

    align_center = Alignment(
        horizontal="center", vertical="center", wrap_text=True
    )

    thin_black = Side(border_style="thin", color="000000")
    medium_black = Side(border_style="medium", color="000000")

    border_cell = Border(
        left=medium_black, right=medium_black, top=thin_black, bottom=thin_black
    )
    border_header = Border(
        left=medium_black,
        right=medium_black,
        top=medium_black,
        bottom=medium_black,
    )

    headers = ["DATE", "PORT OF CALL", "TIME\nZONE", "SUNRISE", "SUNSET"]
    ws.row_dimensions[1].height = 28

    for col_idx, h_text in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h_text)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = border_header

    def format_time_val(t_val):
        t_str = str(t_val).strip()
        parts = t_str.split(":")
        if len(parts) >= 2:
            h = int(parts[0])
            m = parts[1]
            return f"{h}:{m}"
        return t_str

    for idx, row in df_clean.iterrows():
        current_row = idx + 2

        raw_date = row.get("DATE", "")
        if isinstance(raw_date, str) and raw_date:
            try:
                dt_obj = datetime.strptime(raw_date, "%Y-%m-%d")
                date_str = f"{dt_obj.day}/{dt_obj.strftime('%b')}/{dt_obj.year}"
            except ValueError:
                date_str = str(raw_date)
        elif isinstance(raw_date, (datetime, pd.Timestamp)):
            date_str = (
                f"{raw_date.day}/{raw_date.strftime('%b')}/{raw_date.year}"
            )
        else:
            date_str = str(raw_date)

        port_val = str(row.get("PORT / WAYPOINT", ""))
        tz_val = str(row.get("TIME ZONE", ""))
        sr_val = format_time_val(row.get("SUNRISE", ""))
        ss_val = format_time_val(row.get("SUNSET", ""))

        row_data = [date_str, port_val, tz_val, sr_val, ss_val]

        port_lower = port_val.lower()
        if "sea" in port_lower or "fun day" in port_lower:
            row_fill = fill_sea_day
        elif "long beach" in port_lower:
            row_fill = fill_long_beach
        else:
            row_fill = fill_white

        ws.row_dimensions[current_row].height = 20

        for col_idx, val in enumerate(row_data, start=1):
            cell = ws.cell(row=current_row, column=col_idx, value=val)
            cell.alignment = align_center
            cell.border = border_cell
            cell.fill = row_fill
            cell.font = font_date if col_idx == 1 else font_body

    col_widths = {"A": 16, "B": 32, "C": 12, "D": 12, "E": 12}
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


# -----------------------------------------------------------------------------
# SIDEBAR - FILE UPLOAD
# -----------------------------------------------------------------------------
st.sidebar.header("📁 Data Upload")

cal_file = st.sidebar.file_uploader(
    "1. Upload Monthly Calendar (Excel)", type=["xlsx", "xls"]
)
route_files = st.sidebar.file_uploader(
    "2. Upload Route Files (CSV)", type=["csv"], accept_multiple_files=True
)

if cal_file and route_files:
    df_cal = pd.read_excel(cal_file)

    cal_date_col = next(
        (
            c
            for c in df_cal.columns
            if "date" in c.lower() or "data" in c.lower()
        ),
        df_cal.columns[0],
    )
    cal_port_col = next(
        (
            c
            for c in df_cal.columns
            if "location" in c.lower() or "port" in c.lower()
        ),
        df_cal.columns[1],
    )
    cal_code_col = next(
        (
            c
            for c in df_cal.columns
            if "port code" in c.lower() or "code" in c.lower()
        ),
        None,
    )
    cal_cruise_col = next(
        (
            c
            for c in df_cal.columns
            if "cruise" in c.lower() or "crociera" in c.lower()
        ),
        None,
    )

    df_cal["Date_Parsed"] = pd.to_datetime(df_cal[cal_date_col]).dt.date
    df_cal["Location_Clean"] = (
        df_cal[cal_port_col].astype(str).str.strip().str.upper()
    )
    df_cal["Port_Code_Clean"] = (
        df_cal[cal_code_col].astype(str).str.strip().str.upper()
        if cal_code_col
        else ""
    )

    df_cal = df_cal.sort_values(by="Date_Parsed").reset_index(drop=True)

    st.subheader("🔍 Sequential Route Analysis")

    results_from_routes = []

    for rf in route_files:
        try:
            df_r = pd.read_csv(rf)

            time_col = next(
                (
                    c
                    for c in df_r.columns
                    if "arrival time" in c.lower() or "time" in c.lower()
                ),
                None,
            )
            lat_col = next(
                (c for c in df_r.columns if "lat" in c.lower()), None
            )
            lon_col = next(
                (c for c in df_r.columns if "lon" in c.lower()), None
            )

            if not time_col or not lat_col or not lon_col:
                st.info(
                    f"ℹ️ File `{rf.name}` skipped: missing required time or coordinate columns."
                )
                continue

            name_cols = [
                c
                for c in df_r.columns
                if any(k in c.lower() for k in ["name", "port", "location"])
                and not any(k in c.lower() for k in ["no", "num", "id"])
            ]
            name_col = (
                name_cols[0]
                if name_cols
                else next(
                    (c for c in df_r.columns if "waypoint" in c.lower()), None
                )
            )

            tz_col = next(
                (
                    c
                    for c in df_r.columns
                    if any(
                        k in c.lower()
                        for k in [
                            "utc offset",
                            "time zone",
                            "timezone",
                            "offset",
                            "tz",
                        ]
                    )
                ),
                None,
            )

            df_r["dt_parsed"] = pd.to_datetime(df_r[time_col], dayfirst=True)
            df_r["date_only"] = df_r["dt_parsed"].dt.date
            min_date = df_r["date_only"].min()
            df_r["rel_day"] = df_r["date_only"].apply(
                lambda d: (d - min_date).days if pd.notnull(d) else 0
            )

            route_duration_days = int(
                df_r["rel_day"].max() - df_r["rel_day"].min()
            )

            filename_port_codes = re.findall(r"\b[A-Z]{3}\b", rf.name.upper())

            start_code = (
                filename_port_codes[0] if len(filename_port_codes) >= 1 else ""
            )
            end_code = (
                filename_port_codes[1] if len(filename_port_codes) >= 2 else ""
            )

            start_wp = df_r.iloc[0]
            end_wp = df_r.iloc[-1]

            start_name_csv = (
                clean_text_val(start_wp[name_col]).upper() if name_col else ""
            )
            end_name_csv = (
                clean_text_val(end_wp[name_col]).upper() if name_col else ""
            )

            # Aggiunge i waypoints del CSV al database dinamico dei porti
            for _, r_row in df_r.iterrows():
                wp_n = clean_text_val(r_row[name_col]).upper() if name_col else ""
                if wp_n and wp_n not in PORT_DB:
                    w_lat = parse_coordinate(r_row[lat_col])
                    w_lon = parse_coordinate(r_row[lon_col])
                    w_tz = str(r_row[tz_col]) if tz_col else "7 W"
                    if w_lat != 0.0 or w_lon != 0.0:
                        PORT_DB[wp_n] = {"lat": w_lat, "lon": w_lon, "tz": w_tz}

            st.write(
                f"📄 **Route CSV:** `{rf.name}` | "
                f"**Origin:** `{start_code or start_name_csv}` $\\rightarrow$ **Destination:** `{end_code or end_name_csv}` | "
                f"**Route Duration:** {route_duration_days + 1} days"
            )

            matched_sequence_dates = []

            for i in range(len(df_cal)):
                row_start = df_cal.iloc[i]
                date_start = row_start["Date_Parsed"]

                start_match = (
                    (start_code and start_code == row_start["Port_Code_Clean"])
                    or (
                        start_code and start_code in row_start["Location_Clean"]
                    )
                    or (
                        start_name_csv
                        and start_name_csv in row_start["Location_Clean"]
                    )
                )

                if start_match:
                    target_end_date = date_start + timedelta(
                        days=route_duration_days
                    )

                    matching_end_rows = df_cal[
                        df_cal["Date_Parsed"] == target_end_date
                    ]

                    for _, row_end in matching_end_rows.iterrows():
                        end_match = (
                            (
                                end_code
                                and end_code == row_end["Port_Code_Clean"]
                            )
                            or (
                                end_code
                                and end_code in row_end["Location_Clean"]
                            )
                            or (
                                end_name_csv
                                and end_name_csv in row_end["Location_Clean"]
                            )
                        )

                        if end_match:
                            cruise_id = (
                                str(row_start[cal_cruise_col])
                                if cal_cruise_col
                                and pd.notna(row_start[cal_cruise_col])
                                else f"PTP_{date_start}"
                            )
                            matched_sequence_dates.append(
                                (cruise_id, date_start)
                            )

            if matched_sequence_dates:
                st.success(
                    f"✅ Matching sequence found for dates: {[d.strftime('%Y-%m-%d') for _, d in matched_sequence_dates]}"
                )

            for c_code, start_date in matched_sequence_dates:
                for rel_day, day_group in df_r.groupby("rel_day"):
                    actual_date = start_date + timedelta(days=int(rel_day))
                    mid_point = day_group.iloc[len(day_group) // 2]

                    lat = parse_coordinate(mid_point[lat_col])
                    lon = parse_coordinate(mid_point[lon_col])

                    tz_raw = mid_point[tz_col] if tz_col else "0"
                    tz_offset = parse_tz_offset(tz_raw)
                    tz_str_formatted = format_tz_string(tz_raw, tz_offset)

                    results_from_routes.append({
                        "CRUISE CODE": c_code,
                        "DATE": actual_date.strftime("%Y-%m-%d"),
                        "DATE_OBJ": actual_date,
                        "LATITUDE": lat,
                        "LONGITUDE": lon,
                        "TIME ZONE": tz_str_formatted,
                        "TZ_OFFSET": tz_offset,
                        "DT_PARSED": mid_point["dt_parsed"],
                    })

        except Exception as e:
            st.warning(f"⚠️ Could not process file `{rf.name}`: {e}")

    # -----------------------------------------------------------------------------
    # MASTER SCHEDULE CONSTRUCTION & TWO-PASS INTERPOLATION
    # -----------------------------------------------------------------------------
    if not df_cal.empty:
        df_routes = (
            pd.DataFrame(results_from_routes)
            if results_from_routes
            else pd.DataFrame()
        )

        master_schedule = []

        # PASSAGGIO 1: Mappatura iniziale con Porti Noti e file CSV
        for idx, cal_row in df_cal.iterrows():
            c_date = cal_row["Date_Parsed"]
            date_str = c_date.strftime("%Y-%m-%d")
            port_name = clean_text_val(cal_row[cal_port_col])
            cruise_id = (
                str(cal_row[cal_cruise_col])
                if cal_cruise_col and pd.notna(cal_row[cal_cruise_col])
                else "CRUISE"
            )

            lat, lon, tz_str, tz_offset = None, None, None, None

            # 1a. Cerca nei dati ufficiali derivati dai file CSV caricati
            matched_res = None
            if not df_routes.empty and "DATE" in df_routes.columns:
                match_df = df_routes[df_routes["DATE"] == date_str]
                if not match_df.empty:
                    matched_res = match_df.iloc[0].to_dict()

            if matched_res and matched_res.get("LATITUDE") is not None:
                lat = matched_res["LATITUDE"]
                lon = matched_res["LONGITUDE"]
                tz_str = matched_res["TIME ZONE"]
                tz_offset = matched_res["TZ_OFFSET"]
            else:
                # 1b. Cerca nel database dei porti tramite corrispondenza testuale
                coords = lookup_port_coords(port_name, PORT_DB)
                if coords:
                    lat, lon, tz_str = coords
                    tz_offset = parse_tz_offset(tz_str)

            master_schedule.append({
                "CRUISE CODE": cruise_id,
                "DATE_OBJ": c_date,
                "DATE": date_str,
                "PORT / WAYPOINT": port_name,
                "LATITUDE": lat,
                "LONGITUDE": lon,
                "TIME ZONE": tz_str,
                "TZ_OFFSET": tz_offset,
            })

        # PASSAGGIO 2: Interpolazione Lineare degli Intervalli (Sea Days)
        n_items = len(master_schedule)
        known_indices = [
            i
            for i, item in enumerate(master_schedule)
            if item["LATITUDE"] is not None
        ]

        if known_indices:
            # Piattaforma i valori iniziali se il primo giorno non è noto
            first_k = known_indices[0]
            for i in range(0, first_k):
                master_schedule[i]["LATITUDE"] = master_schedule[first_k][
                    "LATITUDE"
                ]
                master_schedule[i]["LONGITUDE"] = master_schedule[first_k][
                    "LONGITUDE"
                ]
                master_schedule[i]["TIME ZONE"] = master_schedule[first_k][
                    "TIME ZONE"
                ]
                master_schedule[i]["TZ_OFFSET"] = master_schedule[first_k][
                    "TZ_OFFSET"
                ]

            # Piattaforma i valori finali se gli ultimi giorni non sono noti
            last_k = known_indices[-1]
            for i in range(last_k + 1, n_items):
                master_schedule[i]["LATITUDE"] = master_schedule[last_k][
                    "LATITUDE"
                ]
                master_schedule[i]["LONGITUDE"] = master_schedule[last_k][
                    "LONGITUDE"
                ]
                master_schedule[i]["TIME ZONE"] = master_schedule[last_k][
                    "TIME ZONE"
                ]
                master_schedule[i]["TZ_OFFSET"] = master_schedule[last_k][
                    "TZ_OFFSET"
                ]

            # Interpolazione progressiva tra ogni coppia di porti noti adiacenti
            for k in range(len(known_indices) - 1):
                idx1 = known_indices[k]
                idx2 = known_indices[k + 1]

                if idx2 - idx1 > 1:  # Presenza di uno o più Sea Days intermedi
                    p1 = master_schedule[idx1]
                    p2 = master_schedule[idx2]

                    total_steps = idx2 - idx1
                    for step in range(1, total_steps):
                        curr_idx = idx1 + step
                        ratio = step / total_steps

                        lat_interp = p1["LATITUDE"] + ratio * (
                            p2["LATITUDE"] - p1["LATITUDE"]
                        )
                        lon_interp = p1["LONGITUDE"] + ratio * (
                            p2["LONGITUDE"] - p1["LONGITUDE"]
                        )

                        tz_off1 = (
                            p1["TZ_OFFSET"]
                            if p1["TZ_OFFSET"] is not None
                            else 0.0
                        )
                        tz_off2 = (
                            p2["TZ_OFFSET"]
                            if p2["TZ_OFFSET"] is not None
                            else 0.0
                        )
                        tz_off_interp = tz_off1 + ratio * (tz_off2 - tz_off1)

                        master_schedule[curr_idx]["LATITUDE"] = lat_interp
                        master_schedule[curr_idx]["LONGITUDE"] = lon_interp
                        master_schedule[curr_idx]["TZ_OFFSET"] = tz_off_interp
                        master_schedule[curr_idx]["TIME ZONE"] = (
                            format_tz_string(p1["TIME ZONE"], tz_off_interp)
                        )

        # PASSAGGIO 3: Calcolo delle effemeridi (Alba e Tramonto)
        final_rows = []
        for item in master_schedule:
            lat = item["LATITUDE"]
            lon = item["LONGITUDE"]
            tz_off = (
                item["TZ_OFFSET"] if item["TZ_OFFSET"] is not None else 0.0
            )

            if lat is not None and lon is not None:
                sr, ss = calculate_sun_events_native(
                    lat, lon, item["DATE_OBJ"], tz_off
                )
            else:
                sr, ss = "N/A", "N/A"

            final_rows.append({
                "CRUISE CODE": item["CRUISE CODE"],
                "DATE": item["DATE"],
                "PORT / WAYPOINT": item["PORT / WAYPOINT"],
                "TIME ZONE": item["TIME ZONE"],
                "SUNRISE": sr,
                "SUNSET": ss,
                "LATITUDE": lat,
                "LONGITUDE": lon,
            })

        df_out = pd.DataFrame(final_rows)

        # Deduplicazione per singola Data Solare
        df_out = df_out.drop_duplicates(subset=["DATE"], keep="first")
        df_out = df_out.sort_values(by="DATE").reset_index(drop=True)

        st.markdown("---")
        st.subheader("📊 Automatic Ephemeris Calculation Results")
        st.dataframe(df_out, use_container_width=True)

        excel_data = export_styled_excel(df_out)

        st.download_button(
            label="📥 Download Complete Excel Report",
            data=excel_data,
            file_name=f'Naval_Ephemeris_{datetime.now().strftime("%Y%m%d")}.xlsx',
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
else:
    st.info(
        "Please upload the Excel calendar and route CSV files to start automatic mapping."
    )
